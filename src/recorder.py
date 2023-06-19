"""Record training statistics"""
import logging
from contextlib import contextmanager
from pathlib import Path
from typing import Tuple

import openpyxl
import tensorflow as tf
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from load_dataset import create_class_filter, get_test_dataset
from tracker import TrainInfo, TrainType
from visualization import plot_summary


class TrainStatistics:
    """Record evaluation results"""

    def __init__(self, export_path: Path):
        """Create or load statistic file

        Parameters
        ----------
        export_path: Path
            Folder to save excel report
        """
        self.report_path = Path("reports")
        self.logger = logging.getLogger("raido.report")
        self.export_folder = export_path
        self.report_path = (export_path / "statistics.xlsx").absolute()
        self.report_path.parent.mkdir(parents=True, exist_ok=True)
        if not self.report_path.exists():
            self.logger.info("Workbook not exist. Creating a new")
            self._create_new()

    def save_train_history(
        self, train_info: TrainInfo, history: tf.keras.callbacks.History
    ) -> None:
        """Record model training information and create plots

        Parameters
        ----------
        train_info: TrainInfo
            Model training information
        history: tf.keras.callbacks.History
            Train history
        """
        self.logger.info("Recording new train metrics for %s", train_info)
        with self._get_workbook() as workbook:
            sheets = (
                workbook["Accuracy"],
                workbook["Loss"],
                workbook["ValidationAccuracy"],
                workbook["ValidationLoss"],
            )
            metrics = ("accuracy", "loss", "val_accuracy", "val_loss")
            for sheet, metric_name in zip(sheets, metrics):
                row = self._get_metric_cell(sheet, train_info)
                sheet[row] = history.history[metric_name][-1]

    def save_evaluation(self, model: tf.keras.Model, train_info: TrainInfo) -> None:
        """Save evaluation metrics on test dataset

        Parameters
        ----------
        model: tf.keras.Model
            Model to evaluate
        train_info: TrainInfo
            Train information
        """
        test_ds = get_test_dataset()
        self.logger.info("Recording evaluation metrics for %s", train_info)
        with self._get_workbook() as workbook:
            loss_cells = self._get_class_cell(workbook["LossPerClass"], train_info)
            acc_cells = self._get_class_cell(workbook["AccuracyPerClass"], train_info)
            for class_index, (a_cell, l_cell) in enumerate(zip(acc_cells, loss_cells)):
                if class_index < 3:
                    _class_ds = test_ds.filter(create_class_filter(class_index)).batch(
                        1
                    )
                    loss, accuracy = model.evaluate(_class_ds)
                else:
                    loss, accuracy = model.evaluate(test_ds.batch(16))
                workbook["AccuracyPerClass"][a_cell] = accuracy
                workbook["LossPerClass"][l_cell] = loss

    def plot_train_summary(self):
        """Save summary plots"""
        with self._get_workbook() as workbook:
            sheets = (
                (workbook["Accuracy"], "Train accuracy"),
                (workbook["Loss"], "Train loss"),
                (workbook["ValidationAccuracy"], "Validation accuracy"),
                (workbook["ValidationLoss"], "Validation loss"),
            )
            for sheet, title in sheets:
                self.logger.info("Plotting %s", title)
                x_values = []
                y_values = []
                z_values = [[], [], []]
                titles = tuple(
                    sheet[f"{get_column_letter(offset)}1"].value for offset in (3, 4, 5)
                )
                for row in sheet.iter_rows(
                    min_row=2, min_col=1, max_col=5, values_only=True
                ):
                    try:
                        x_val = float(row[0])
                        y_val = float(row[1])
                    except (ValueError, TypeError):
                        self.logger.warning("Got error on row %s", row)
                        continue
                    x_values.append(x_val)
                    y_values.append(y_val)
                    for index, offset in enumerate([2, 3, 4]):
                        try:
                            z_values[index].append(float(row[offset]))
                        except (ValueError, TypeError):
                            self.logger.warning(
                                "Got error on row %s with offset %s", row, offset
                            )
                            z_values[index].append(0)
                            continue
                plot_summary(
                    x_axis=x_values,
                    y_axis=y_values,
                    values=z_values,
                    titles=titles,
                    invert_bar="loss" in title,
                    export_path=self.export_folder / f"{title}.png",
                )

    def plot_class_summary(self) -> None:
        """Save class performance plots"""
        with self._get_workbook() as workbook:
            sheets = (
                (workbook["AccuracyPerClass"], "Accuracy on test dataset"),
                (workbook["LossPerClass"], "Loss on test dataset"),
            )
            titles = tuple(
                workbook["LossPerClass"][f"{get_column_letter(offset)}2"].value
                for offset in (3, 4, 5, 6)
            )
            for sheet, title in sheets:
                for shift, _type in enumerate(TrainType):
                    self.logger.info("Plotting %s type %s", title, _type.value)
                    x_values = []
                    y_values = []
                    z_values = [[], [], [], []]
                    for row in sheet.iter_rows(
                        min_row=3, min_col=1, max_col=14, values_only=True
                    ):
                        try:
                            x_val = float(row[0])
                            y_val = float(row[1])
                        except (ValueError, TypeError):
                            self.logger.warning("Got error on row %s", row)
                            continue
                        x_values.append(x_val)
                        y_values.append(y_val)
                        for index, offset in enumerate(
                            range(2 + 4 * shift, 6 + 4 * shift)
                        ):
                            try:
                                z_values[index].append(float(row[offset]))
                            except (ValueError, TypeError):
                                self.logger.warning(
                                    "Got error on row %s with offset %s", row, offset
                                )
                                z_values[index].append(0)
                                continue
                    plot_summary(
                        x_axis=x_values,
                        y_axis=y_values,
                        values=z_values,
                        titles=titles,
                        invert_bar="loss" in title,
                        export_path=self.export_folder / f"{title} ({_type.value}).png",
                    )

    @contextmanager
    def _get_workbook(self) -> openpyxl.Workbook:
        """Create or open XLS file

        Returns
        -------
        openpyxl.Workbook
            New or existing workbook
        """
        self.logger.info("Loading workbook")
        workbook = openpyxl.load_workbook(filename=self.report_path, data_only=True)
        yield workbook
        self.logger.info("Saving workbook")
        workbook.save(self.report_path)

    def _create_new(self) -> None:
        """Create new workbook"""
        classes_names = get_test_dataset().class_names + [
            "total",
        ]
        new_wb = openpyxl.Workbook()
        new_wb.create_sheet("Accuracy")
        new_wb.create_sheet("Loss")
        new_wb.create_sheet("ValidationAccuracy")
        new_wb.create_sheet("ValidationLoss")
        for sheet in (
            new_wb["Accuracy"],
            new_wb["Loss"],
            new_wb["ValidationAccuracy"],
            new_wb["ValidationLoss"],
        ):
            sheet["A1"] = "Reduction"
            sheet["B1"] = "Corruption"
            for offset, train_type in enumerate(TrainType, start=3):
                sheet[f"{get_column_letter(offset)}1"] = f"{train_type}".capitalize()
        new_wb.create_sheet("AccuracyPerClass")
        new_wb.create_sheet("LossPerClass")
        for sheet in (new_wb["AccuracyPerClass"], new_wb["LossPerClass"]):
            sheet.merge_cells("A1:A2")
            cell = sheet.cell(row=1, column=1)
            cell.value = "Reduction"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.merge_cells("B1:B2")
            cell = sheet.cell(row=1, column=2)
            cell.value = "Corruption"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for col_offset, train_type in enumerate(TrainType):
                cell_index = len(classes_names) * col_offset + 3
                sheet.merge_cells(
                    f"{get_column_letter(cell_index)}1:{get_column_letter(cell_index + len(classes_names) - 1)}1"
                )
                cell = sheet.cell(row=1, column=cell_index)
                cell.value = f"{train_type}".capitalize()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                for class_offset, class_name in enumerate(classes_names):
                    sheet[
                        f"{get_column_letter(cell_index + class_offset)}2"
                    ] = f"{class_name}".capitalize()
        del new_wb["Sheet"]
        new_wb.save(self.report_path)

    def _get_metric_cell(self, sheet: Worksheet, train_info: TrainInfo) -> str:
        """Find first empty point or return existing one according to train_info

        Parameters
        ----------
        sheet: Worksheet
            Worksheet to search
        train_info: TrainInfo
            Train information

        Returns
        -------
        str
            First empty or matched cell
        """
        offset_col = 2 + list(_type for _type in TrainType).index(train_info.train_type)
        row_index = 1
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=row_index, min_col=1, max_col=5), start=row_index
        ):
            if (
                row[0].value == train_info.reduction
                and row[1].value == train_info.corruption
            ):
                self.logger.debug(
                    "Cell found at coordinate %s", row[offset_col].coordinate
                )
                return row[offset_col].coordinate
        sheet[f"A{row_index+1}"] = train_info.reduction
        sheet[f"B{row_index+1}"] = train_info.corruption
        cell_coord = f"{get_column_letter(offset_col)}{row_index+1}"
        self.logger.info("Cell not found.Appending %s", cell_coord)
        return cell_coord

    def _get_class_cell(
        self, sheet: Worksheet, train_info: TrainInfo
    ) -> Tuple[str, str, str]:
        """Find first empty point or return existing one according to train_info

        Parameters
        ----------
        sheet: Worksheet
            Worksheet to search
        train_info: TrainInfo
            Train information

        Returns
        -------
        Tuple[str, str, str, str]
            First empty or matched cells (for all classes)
        """
        offset_col = (
            2 + list(_type for _type in TrainType).index(train_info.train_type) * 3
        )
        row_index = 1
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=row_index, min_col=1), start=row_index
        ):
            self.logger.debug(
                "Testing cell at coordinate %s", row[offset_col].coordinate
            )
            if (
                row[0].value == train_info.reduction
                and row[1].value == train_info.corruption
            ):
                self.logger.debug(
                    "Cell found at coordinate %s", row[offset_col].coordinate
                )
                return (
                    row[offset_col].coordinate,
                    row[offset_col + 1].coordinate,
                    row[offset_col + 2].coordinate,
                    row[offset_col + 3].coordinate,
                )
        sheet[f"A{row_index+1}"] = train_info.reduction
        sheet[f"B{row_index+1}"] = train_info.corruption
        cell_coord = (
            f"{get_column_letter(offset_col+1)}{row_index+1}",
            f"{get_column_letter(offset_col+2)}{row_index+1}",
            f"{get_column_letter(offset_col+3)}{row_index+1}",
            f"{get_column_letter(offset_col+4)}{row_index+1}",
        )
        self.logger.info("Cell not found.Appending %s", cell_coord)
        return cell_coord
