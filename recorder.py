"""Record training statistics"""
import logging
from pathlib import Path

import numpy as np
import openpyxl
import tensorflow as tf
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter
from matplotlib import pyplot as plt
from matplotlib import use


from config import TrainInfo, TrainType, PlotSettings

from load_dataset import get_test_dataset, create_class_filter
from visualization import (
    plot_history,
    plot_confusion_matrix,
    plot_test_dataset_evaluation,
)


class TrainStatistics:
    """Record evaluation results"""

    def __init__(self, network: str):
        self.report_path = Path("reports")
        self.logger = logging.getLogger("raido.dataset")
        self.report_path = Path("reports", network, "statistics.xlsx")

    def add_train_metrics(
        self, train_info: TrainInfo, history: tf.keras.callbacks.History
    ) -> None:
        """Record model training information and create plots

         Parameters
         ----------
         train_info: TrainInfo
             Model training information
        history: tf.keras.callbacks.History
             Train history
        """
        self.logger.info("Recording new train metrics for %s", train_info)
        workbook = self._get_workbook()
        row = self._get_metric_emtpy_row(workbook)
        sheets = (
            workbook["Accuracy"],
            workbook["Loss"],
            workbook["ValidationAccuracy"],
            workbook["ValidationLoss"],
        )
        metrics = ("accuracy", "loss", "val_accuracy", "val_loss")
        train_type = list(_type for _type in TrainType)
        for sheet, metric_name in zip(sheets, metrics):
            sheet[f"A{row}"] = train_info.reduction
            sheet[f"B{row}"] = train_info.corruption
            sheet[
                f"{get_column_letter(train_type.index(train_info.train_type) + 3)}{row}"
            ] = history.history[metric_name][-1]
        self.logger.info("Saving workbook")
        workbook.save(self.report_path)
        self.logger.info("Plotting train history")
        plot_history(history, train_info)

    def evaluate_point(
        self,
        train_info: TrainInfo,
        model: tf.keras.models.Model,
        dataset: tf.data.Dataset,
    ) -> None:
        """Evaluate model on validation and test dataset

        Parameters
        ----------
        train_info: TrainInfo
            Model train information
        model: tf.keras.models.Model
            Trained model
        dataset: tf.data.Dataset
            Validation dataset for plots
        """
        self.logger.info("Evaluating model performance - %s", train_info)
        test_dataset = get_test_dataset()
        y_true = []
        y_pred = []
        labels = test_dataset.class_names
        for image, class_tensor in iter(dataset):
            y_true.append(np.argmax(class_tensor.numpy()))
            y_pred.append(np.argmax(model.predict(tf.data.Dataset.from_tensors(image).batch(1), verbose=0), axis=1))
        self.logger.info("Plotting confusion matrix on validation dataset")
        plot_confusion_matrix(y_true, y_pred, labels, train_info, "validation")

        y_true = []
        y_pred = []
        for image, class_tensor in iter(test_dataset):
            y_true.append(np.argmax(class_tensor.numpy()))
            y_pred.append(np.argmax(model.predict(tf.data.Dataset.from_tensors(image).batch(1), verbose=0), axis=1))
        self.logger.info("Plotting confusion matrix on test dataset")
        plot_confusion_matrix(y_true, y_pred, labels, train_info, "test")

        self.logger.info("Saving evaluation data")
        workbook = self._get_workbook()
        insert_row = self._get_class_emtpy_row(workbook)
        type_offset = (
            list(_type for _type in TrainType).index(train_info.train_type)
            * len(labels)
            + 3
        )
        for class_index in range(len(labels)):
            loss, accuracy = model.evaluate(
                test_dataset.filter(create_class_filter(class_index)).batch(1),
                verbose=2,
            )
            insert_index = f"{get_column_letter(type_offset + class_index)}{insert_row}"
            workbook["AccuracyPerClass"][insert_index] = accuracy
            workbook["LossPerClass"][insert_index] = loss
        self.logger.info("Saving workbook")
        workbook.save(self.report_path)
        self.logger.info("Platting dataset evaluation graphs")
        plot_test_dataset_evaluation(model, train_info)

    def _get_workbook(self) -> openpyxl.Workbook:
        """Create or open XLS file

        Returns
        -------
        openpyxl.Workbook
            New or existing workbook
        """
        if self.report_path.exists():
            self.logger.info("Loading workbook")
            return openpyxl.load_workbook(filename=self.report_path, data_only=True)
        self.logger.info("Workbook not exist. Creating a new")
        new_wb = openpyxl.Workbook()
        new_wb.create_sheet("Accuracy")
        new_wb.create_sheet("Loss")
        new_wb.create_sheet("ValidationAccuracy")
        new_wb.create_sheet("ValidationLoss")
        for sheet in (
            new_wb["Accuracy"],
            new_wb["Loss"],
            new_wb["ValidationAccuracy"],
            new_wb["ValidationLoss"],
        ):
            sheet["A1"] = "Reduction"
            sheet["B1"] = "Corruption"
            for offset, train_type in enumerate(TrainType, start=3):
                sheet[f"{get_column_letter(offset)}1"] = f"{train_type}".capitalize()
        new_wb.create_sheet("AccuracyPerClass")
        new_wb.create_sheet("LossPerClass")
        classes = get_test_dataset().class_names
        for sheet in (new_wb["AccuracyPerClass"], new_wb["LossPerClass"]):
            sheet.merge_cells("A1:A2")
            cell = sheet.cell(row=1, column=1)
            cell.value = "Reduction"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.merge_cells("B1:B2")
            cell = sheet.cell(row=1, column=2)
            cell.value = "Corruption"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for col_offset, train_type in enumerate(TrainType):
                cell_index = len(classes) * col_offset + 3
                sheet.merge_cells(
                    f"{get_column_letter(cell_index)}1:{get_column_letter(cell_index + len(classes) - 1)}1"
                )
                cell = sheet.cell(row=1, column=cell_index)
                cell.value = f"{train_type}".capitalize()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                for class_offset, class_name in enumerate(classes):
                    sheet[
                        f"{get_column_letter(cell_index + class_offset)}2"
                    ] = f"{class_name}".capitalize()
        del new_wb["Sheet"]
        new_wb.save(self.report_path)
        return new_wb

    def is_already_trained(self, train_info: TrainInfo) -> bool:
        """Return if point already recorder"""
        workbook = self._get_workbook()
        worksheet = workbook["Accuracy"]
        for row in range(1, worksheet.max_row):
            if (
                worksheet.cell(row, 1).value == train_info.reduction
                and worksheet.cell(row, 2).value == train_info.corruption
            ):
                self.logger.info("Point %s already trained", train_info)
                return True
            if worksheet.cell(row, 1).value is None:
                return False
        return False

    def plot_general_summary(self, train_info: TrainInfo) -> None:
        """Plot statistics graphs

        Parameters
        ----------
        train_info: TrainInfo
            Model training info

        """
        plot_settings = PlotSettings()
        workbook = self._get_workbook()
        # use("Agg")
        # fig = plt.figure(figsize=plot_settings.plot_size, dpi=plot_settings.plot_dpi)
        # fig.tight_layout()
        # subfigs = fig.subfigures(nrows=2, ncols=1)

    @staticmethod
    def _get_metric_emtpy_row(workbook: openpyxl.Workbook) -> int:
        """Get first empty row

        Returns
        -------
        int
            Empty row index
        """
        worksheet = workbook["Accuracy"]
        for row in range(1, worksheet.max_row):
            if worksheet.cell(row, 1).value is None:
                return row
        return worksheet.max_row

    @staticmethod
    def _get_class_emtpy_row(workbook: openpyxl.Workbook) -> int:
        """Get first empty row

        Returns
        -------
        int
            Empty row index
        """
        worksheet = workbook["AccuracyPerClass"]
        for row in range(1, worksheet.max_row):
            if worksheet.cell(row, 1).value is None:
                return row
        return worksheet.max_row
